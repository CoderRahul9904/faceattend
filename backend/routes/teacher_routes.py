import io
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, Depends, HTTPException, status, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, timezone, date as date_type

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from database import get_db
import models
import schemas
import auth
import face_utils

router = APIRouter(tags=["teacher"])

# Subject Routes

@router.post("/subjects/", response_model=schemas.SubjectResponse, status_code=status.HTTP_201_CREATED)
def create_subject(
    subject_in: schemas.SubjectCreate,
    current_user: models.User = Depends(auth.require_role(["teacher"])),
    db: Session = Depends(get_db)
):
    # Check if subject code already exists
    existing = db.query(models.Subject).filter(models.Subject.code == subject_in.code.strip()).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Subject code already exists"
        )
        
    new_subject = models.Subject(
        name=subject_in.name.strip(),
        code=subject_in.code.strip(),
        teacher_id=current_user.id
    )
    db.add(new_subject)
    db.commit()
    db.refresh(new_subject)
    return new_subject

@router.get("/subjects/", response_model=list[schemas.SubjectResponse])
def list_subjects(
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    # If the user is a teacher, show their subjects. If student, show all subjects.
    if current_user.role == "teacher":
        return db.query(models.Subject).filter(models.Subject.teacher_id == current_user.id).all()
    return db.query(models.Subject).all()

@router.get("/students", response_model=list[schemas.UserResponse])
def list_students(
    current_user: models.User = Depends(auth.require_role(["teacher"])),
    db: Session = Depends(get_db)
):
    return db.query(models.User).filter(models.User.role == "student").all()

# Session Routes

@router.post("/sessions/start")
def start_session(
    session_start: schemas.SessionStart,
    current_user: models.User = Depends(auth.require_role(["teacher"])),
    db: Session = Depends(get_db)
):
    # Verify subject exists and belongs to the teacher
    subject = db.query(models.Subject).filter(
        models.Subject.id == session_start.subject_id,
        models.Subject.teacher_id == current_user.id
    ).first()
    if not subject:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Subject not found or does not belong to you"
        )
        
    # Check if there is already an active session for this teacher
    active_session = db.query(models.Session).filter(
        models.Session.teacher_id == current_user.id,
        models.Session.status == "active"
    ).first()
    if active_session:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="An active session already exists for this teacher"
        )
        
    new_session = models.Session(
        subject_id=session_start.subject_id,
        teacher_id=current_user.id,
        date=datetime.now(timezone.utc).date(),
        start_time=datetime.now(timezone.utc),
        scan_interval_seconds=session_start.scan_interval_seconds,
        status="active"
    )
    db.add(new_session)
    db.commit()
    db.refresh(new_session)
    return {"session_id": new_session.id}

@router.post("/sessions/stop")
def stop_session(
    session_stop: schemas.SessionStop,
    current_user: models.User = Depends(auth.require_role(["teacher"])),
    db: Session = Depends(get_db)
):
    # Find active session
    session = db.query(models.Session).filter(
        models.Session.id == session_stop.session_id,
        models.Session.teacher_id == current_user.id
    ).first()
    
    if not session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Session not found"
        )
        
    if session.status == "completed":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Session has already been stopped"
        )
        
    session.end_time = datetime.now(timezone.utc)
    session.status = "completed"
    db.commit()
    
    # Trigger final attendance computation
    face_utils.compute_attendance(session.id, db)
    
    return {"status": "success", "message": "Session stopped and attendance computed"}

@router.get("/sessions/active", response_model=Optional[schemas.SessionResponse])
def get_active_session(
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    # Find any active session. If current user is teacher, check theirs, else check globally active
    query = db.query(models.Session).filter(models.Session.status == "active")
    if current_user.role == "teacher":
        query = query.filter(models.Session.teacher_id == current_user.id)
        
    session = query.first()
    if not session:
        return None
        
    return schemas.SessionResponse(
        id=session.id,
        subject_id=session.subject_id,
        teacher_id=session.teacher_id,
        date=session.date,
        start_time=session.start_time,
        end_time=session.end_time,
        scan_interval_seconds=session.scan_interval_seconds,
        status=session.status,
        subject_name=session.subject.name
    )

@router.get("/sessions/", response_model=list[schemas.SessionResponse])
def list_sessions(
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    # If teacher, list all their sessions. If student, list all sessions
    query = db.query(models.Session)
    if current_user.role == "teacher":
        query = query.filter(models.Session.teacher_id == current_user.id)
        
    sessions = query.order_by(models.Session.start_time.desc()).all()
    
    # Map to schemas including subject name
    response_list = []
    for s in sessions:
        response_list.append(
            schemas.SessionResponse(
                id=s.id,
                subject_id=s.subject_id,
                teacher_id=s.teacher_id,
                date=s.date,
                start_time=s.start_time,
                end_time=s.end_time,
                scan_interval_seconds=s.scan_interval_seconds,
                status=s.status,
                subject_name=s.subject.name
            )
        )
    return response_list

# Live Scan Endpoint

@router.post("/sessions/{session_id}/scan")
async def scan_session_frame(
    session_id: int,
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth.require_role(["teacher"])),
    db: Session = Depends(get_db)
):
    # Verify session exists
    session = db.query(models.Session).filter(models.Session.id == session_id).first()
    if not session:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Session not found"
        )
        
    # Verify owner
    if session.teacher_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only the teacher who owns the session can trigger a scan"
        )
        
    # Verify session status is active
    if session.status != "active":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Session is not active"
        )
        
    # Read the image frame bytes
    try:
        frame_bytes = await file.read()
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Could not read frame: {str(e)}"
        )
        
    # Load all face encodings
    known_encodings = face_utils.load_all_encodings(db)
    
    # Find matches
    matches = face_utils.find_matches(frame_bytes, known_encodings)
    matches_dict = {user_id: confidence for user_id, confidence in matches}
    
    # Determine scan number (scan_number = last_scan_number + 1)
    last_scan = db.query(models.Scan).filter(
        models.Scan.session_id == session_id
    ).order_by(models.Scan.scan_number.desc()).first()
    scan_number = (last_scan.scan_number + 1) if last_scan else 1
    
    # Create Scan record
    new_scan = models.Scan(
        session_id=session_id,
        scan_number=scan_number,
        scanned_at=datetime.now(timezone.utc)
    )
    db.add(new_scan)
    db.flush()  # Populates new_scan.id
    
    # Create ScanResult records for every registered student
    students = db.query(models.User).filter(models.User.role == "student").all()
    results_list = []
    
    for student in students:
        detected = student.id in matches_dict
        confidence = matches_dict[student.id] if detected else 0.0
        
        scan_result = models.ScanResult(
            scan_id=new_scan.id,
            student_id=student.id,
            detected=detected,
            confidence=confidence
        )
        db.add(scan_result)
        
        results_list.append({
            "student_id": student.student_id,
            "student_name": student.name,
            "detected": detected,
            "confidence": confidence
        })
        
    db.commit()
    return results_list


# ──────────────────────────────────────────────────────────────────────────────
# MISSION 5 — Attendance & Export Routes
# ──────────────────────────────────────────────────────────────────────────────

@router.get("/attendance/session/{session_id}", response_model=List[Dict[str, Any]])
def get_session_attendance(
    session_id: int,
    current_user: models.User = Depends(auth.require_role(["teacher"])),
    db: Session = Depends(get_db),
):
    """
    Returns all students with their final present/absent status for a session.
    Teacher must own the session.

    Edge cases:
    - Session not found → 404
    - Session still active → 400 (attendance not yet computed)
    """
    session = db.query(models.Session).filter(models.Session.id == session_id).first()
    if not session:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Session not found")

    if session.teacher_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not own this session",
        )

    if session.status == "active":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Session is still active. Stop the session first to compute attendance.",
        )

    # Fetch attendance records for this session, joined with student
    attendance_records = (
        db.query(models.Attendance)
        .join(models.User, models.Attendance.student_id == models.User.id)
        .filter(models.Attendance.session_id == session_id)
        .all()
    )

    # Count total scans for this session (used for scans_detected / total_scans)
    total_scans = db.query(models.Scan).filter(models.Scan.session_id == session_id).count()

    results = []
    for att in attendance_records:
        student = att.student

        # Count how many scans this student was detected in
        scans_detected = (
            db.query(models.ScanResult)
            .join(models.Scan, models.ScanResult.scan_id == models.Scan.id)
            .filter(
                models.ScanResult.student_id == student.id,
                models.Scan.session_id == session_id,
                models.ScanResult.detected == True,  # noqa: E712
            )
            .count()
        )

        results.append(
            {
                "student_db_id": student.id,
                "student_id": student.student_id,
                "student_name": student.name,
                "status": att.status,
                "scans_detected": scans_detected,
                "total_scans": total_scans,
            }
        )

    return results


@router.get("/attendance/export/{session_id}")
def export_session_attendance(
    session_id: int,
    current_user: models.User = Depends(auth.require_role(["teacher"])),
    db: Session = Depends(get_db),
):
    """
    Generates and returns a downloadable .xlsx file for a session's attendance.

    Columns: Student ID, Student Name, Subject, Date, Status, Scans Detected, Total Scans
    - Header row: bold text, blue background (#3B5BDB), white font
    - Present rows: light green fill (#C8F7C5)
    - Absent rows: light red fill (#FADADD)
    - Auto-adjusted column widths

    Edge cases:
    - Session not found → 404
    - Session still active → 400
    - Non-owner teacher → 403
    """
    session = db.query(models.Session).filter(models.Session.id == session_id).first()
    if not session:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Session not found")

    if session.teacher_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not own this session",
        )

    if session.status == "active":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Session is still active. Stop the session first to export attendance.",
        )

    subject = session.subject
    session_date = str(session.date)  # YYYY-MM-DD

    # Fetch attendance records
    attendance_records = (
        db.query(models.Attendance)
        .join(models.User, models.Attendance.student_id == models.User.id)
        .filter(models.Attendance.session_id == session_id)
        .order_by(models.User.student_id)
        .all()
    )

    total_scans = db.query(models.Scan).filter(models.Scan.session_id == session_id).count()

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Colour palette
    HEADER_BG = "3B5BDB"       # Blue
    HEADER_FG = "FFFFFF"       # White
    GREEN_FILL = "C8F7C5"      # Light green for Present
    RED_FILL = "FADADD"        # Light red for Absent

    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_BG)
    present_fill = PatternFill(fill_type="solid", fgColor=GREEN_FILL)
    absent_fill = PatternFill(fill_type="solid", fgColor=RED_FILL)

    bold_white_font = Font(bold=True, color=HEADER_FG)

    columns = [
        "Student ID",
        "Student Name",
        "Subject",
        "Date",
        "Status",
        "Scans Detected",
        "Total Scans",
    ]

    # ── Header row ────────────────────────────────────────────────────────────
    ws.append(columns)
    header_row = ws[1]
    for cell in header_row:
        cell.font = bold_white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Data rows ─────────────────────────────────────────────────────────────
    for att in attendance_records:
        student = att.student
        scans_detected = (
            db.query(models.ScanResult)
            .join(models.Scan, models.ScanResult.scan_id == models.Scan.id)
            .filter(
                models.ScanResult.student_id == student.id,
                models.Scan.session_id == session_id,
                models.ScanResult.detected == True,  # noqa: E712
            )
            .count()
        )

        row = [
            student.student_id or str(student.id),
            student.name,
            f"{subject.name} ({subject.code})",
            session_date,
            att.status.capitalize(),
            f"{scans_detected}/{total_scans}",
            total_scans,
        ]
        ws.append(row)

        # Apply row fill based on status
        row_idx = ws.max_row
        fill = present_fill if att.status == "present" else absent_fill
        for col_idx in range(1, len(columns) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    # ── Auto-adjust column widths ──────────────────────────────────────────────
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 4  # Padding

    # ── Stream response ───────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"attendance_{session_id}.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
