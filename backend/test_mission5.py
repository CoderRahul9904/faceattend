"""
test_mission5.py — Tests for Mission 5: Attendance & Export Routes

Tests:
1. Student attendance endpoint returns correct grouped data (multi-session)
2. Teacher session-attendance endpoint returns correct present/absent breakdown
3. Excel export returns a valid .xlsx file with correct headers and row count
4. Export rejected (400) if session is still active
5. Non-owner teacher gets 403 on session export
"""

import os
import io
import pytest
import openpyxl
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

# ── Environment must be set before importing app modules ─────────────────────
os.environ["SECRET_KEY"] = "test_secret_m5"
os.environ["ALGORITHM"] = "HS256"
os.environ["ACCESS_TOKEN_EXPIRE_MINUTES"] = "30"

from database import Base, get_db
from main import app
import models
import face_utils

# ── In-memory test database ───────────────────────────────────────────────────
TEST_DB_URL = "sqlite:///./test_faceattend_m5.db"
engine = create_engine(TEST_DB_URL, connect_args={"check_same_thread": False})
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


def override_get_db():
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()


client = TestClient(app)

# Track real user DB IDs so mocks can match them
student_mapping: dict = {}


def mock_load_all_encodings(db):
    import numpy as np
    return {uid: {"enc": uid} for uid in student_mapping.keys()}


def mock_find_matches(frame_bytes, known_encodings_dict):
    """Deterministic mocked matching based on sentinel byte payloads."""
    matches = []
    for uid, name in student_mapping.items():
        if frame_bytes == b"detect_all":
            matches.append((uid, 0.90))
        elif frame_bytes == b"detect_none":
            pass  # nobody detected
        elif frame_bytes == b"detect_alice" and name == "Alice":
            matches.append((uid, 0.85))
    return matches


# ── Fixture: fresh DB per test ────────────────────────────────────────────────
@pytest.fixture(autouse=True)
def setup_and_teardown(monkeypatch):
    app.dependency_overrides[get_db] = override_get_db
    monkeypatch.setattr(face_utils, "load_all_encodings", mock_load_all_encodings)
    monkeypatch.setattr(face_utils, "find_matches", mock_find_matches)

    Base.metadata.create_all(bind=engine)
    yield
    Base.metadata.drop_all(bind=engine)
    engine.dispose()
    if os.path.exists("test_faceattend_m5.db"):
        try:
            os.remove("test_faceattend_m5.db")
        except Exception:
            pass
    student_mapping.clear()
    app.dependency_overrides.pop(get_db, None)


# ── Helpers ───────────────────────────────────────────────────────────────────
def register_and_login(payload: dict) -> tuple[dict, str]:
    """Register user, login, return (user_json, token)."""
    reg = client.post("/auth/register", json=payload)
    assert reg.status_code == 201, reg.text  # /auth/register returns 201 Created
    login = client.post(
        "/auth/login",
        json={"email": payload["email"], "password": payload["password"]},
    )
    assert login.status_code == 200, login.text
    return reg.json(), login.json()["access_token"]


def auth_headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def build_completed_session(
    t_headers: dict,
    subject_id: int,
    scan_payloads: list[bytes],
) -> int:
    """Start a session, run scans, stop session. Returns session_id."""
    sess_resp = client.post(
        "/sessions/start",
        json={"subject_id": subject_id, "scan_interval_seconds": 15},
        headers=t_headers,
    )
    assert sess_resp.status_code == 200, sess_resp.text
    sess_id = sess_resp.json()["session_id"]

    for payload_bytes in scan_payloads:
        r = client.post(
            f"/sessions/{sess_id}/scan",
            files={"file": ("frame.jpg", payload_bytes, "image/jpeg")},
            headers=t_headers,
        )
        assert r.status_code == 200, r.text

    stop = client.post(
        "/sessions/stop",
        json={"session_id": sess_id},
        headers=t_headers,
    )
    assert stop.status_code == 200, stop.text
    return sess_id


# ─────────────────────────────────────────────────────────────────────────────
# TEST 1 — Student attendance: grouped by subject, multiple sessions
# ─────────────────────────────────────────────────────────────────────────────
def test_student_attendance_grouped_by_subject():
    """
    Setup:
    - 1 teacher, 2 students (Alice, Bob)
    - 2 subjects (Math, Physics)
    - Math: 2 completed sessions — Alice present in both; Bob absent in both
    - Physics: 1 completed session — both present
    Verify Alice's /attendance/student response has:
    - 2 subjects (Math entry with 2 sessions, Physics entry with 1 session)
    - Both Math sessions show status=present for Alice
    """
    # Register teacher
    teacher, t_token = register_and_login(
        {"name": "Prof T", "email": "prof@test.com", "password": "pass", "role": "teacher"}
    )
    t_hdrs = auth_headers(t_token)

    # Register students
    alice_data, alice_token = register_and_login(
        {"name": "Alice", "email": "alice@test.com", "password": "pass",
         "role": "student", "student_id": "STU_A"}
    )
    bob_data, bob_token = register_and_login(
        {"name": "Bob", "email": "bob@test.com", "password": "pass",
         "role": "student", "student_id": "STU_B"}
    )
    alice_id = alice_data["id"]
    bob_id = bob_data["id"]
    student_mapping[alice_id] = "Alice"
    student_mapping[bob_id] = "Bob"

    # Create subjects
    math_id = client.post("/subjects/", json={"name": "Math", "code": "M101"}, headers=t_hdrs).json()["id"]
    phys_id = client.post("/subjects/", json={"name": "Physics", "code": "P101"}, headers=t_hdrs).json()["id"]

    # Math session 1: detect_alice → Alice present, Bob absent (2 consec misses = absent)
    build_completed_session(t_hdrs, math_id, [b"detect_alice", b"detect_alice"])

    # Math session 2: detect_all → both present
    build_completed_session(t_hdrs, math_id, [b"detect_all", b"detect_all"])

    # Physics session: detect_all → both present
    build_completed_session(t_hdrs, phys_id, [b"detect_all", b"detect_all"])

    # ── Assert Alice's view ───────────────────────────────────────────────────
    alice_hdrs = auth_headers(alice_token)
    resp = client.get("/attendance/student", headers=alice_hdrs)
    assert resp.status_code == 200, resp.text
    data = resp.json()

    # Should have exactly 2 subjects
    assert len(data) == 2, f"Expected 2 subjects, got {len(data)}"

    subjects_by_name = {entry["subject_name"]: entry for entry in data}
    assert "Math" in subjects_by_name
    assert "Physics" in subjects_by_name

    math_entry = subjects_by_name["Math"]
    assert len(math_entry["sessions"]) == 2

    # Both Math sessions: Alice present in session 2 (detect_all) — session 1 detect_alice also present
    for sess in math_entry["sessions"]:
        assert sess["status"] == "present", f"Expected Alice present but got {sess['status']}"

    phys_entry = subjects_by_name["Physics"]
    assert len(phys_entry["sessions"]) == 1
    assert phys_entry["sessions"][0]["status"] == "present"

    # ── Assert Bob's view ─────────────────────────────────────────────────────
    bob_hdrs = auth_headers(bob_token)
    resp_bob = client.get("/attendance/student", headers=bob_hdrs)
    assert resp_bob.status_code == 200, resp_bob.text
    bob_data_resp = resp_bob.json()

    subjects_bob = {entry["subject_name"]: entry for entry in bob_data_resp}
    math_bob = subjects_bob.get("Math", {})
    math_sessions_bob = math_bob.get("sessions", [])

    # Math session 1: Bob absent (detect_alice → Bob not detected consecutively)
    statuses = [s["status"] for s in math_sessions_bob]
    assert "absent" in statuses, f"Bob should be absent in at least one Math session, got {statuses}"


def test_student_attendance_empty_list_if_no_records():
    """Student with zero attendance records gets [] not an error."""
    _, s_token = register_and_login(
        {"name": "NewStu", "email": "newstu@test.com", "password": "pass",
         "role": "student", "student_id": "STU_NEW"}
    )
    resp = client.get("/attendance/student", headers=auth_headers(s_token))
    assert resp.status_code == 200
    assert resp.json() == []


def test_student_attendance_requires_student_role():
    """Teachers should be forbidden from /attendance/student."""
    _, t_token = register_and_login(
        {"name": "TeacherOnly", "email": "tonly@test.com", "password": "pass", "role": "teacher"}
    )
    resp = client.get("/attendance/student", headers=auth_headers(t_token))
    assert resp.status_code == 403


# ─────────────────────────────────────────────────────────────────────────────
# TEST 2 — Teacher: session attendance breakdown
# ─────────────────────────────────────────────────────────────────────────────
def test_teacher_session_attendance_breakdown():
    """
    Scenario:
    - 2 students, 1 session
    - Scan 1: detect_alice → Alice present, Bob missing
    - Scan 2: detect_alice → Alice present, Bob missing again (2 consec → absent)
    - Session stopped → attendance computed
    Verify /attendance/session/{id} returns:
    - Alice: present
    - Bob: absent
    - scans_detected / total_scans correct
    """
    teacher, t_token = register_and_login(
        {"name": "Prof S", "email": "profs@test.com", "password": "pass", "role": "teacher"}
    )
    t_hdrs = auth_headers(t_token)

    alice_data, _ = register_and_login(
        {"name": "Alice", "email": "alice2@test.com", "password": "pass",
         "role": "student", "student_id": "STU_A2"}
    )
    bob_data, _ = register_and_login(
        {"name": "Bob", "email": "bob2@test.com", "password": "pass",
         "role": "student", "student_id": "STU_B2"}
    )
    student_mapping[alice_data["id"]] = "Alice"
    student_mapping[bob_data["id"]] = "Bob"

    math_id = client.post("/subjects/", json={"name": "Math", "code": "M102"}, headers=t_hdrs).json()["id"]
    sess_id = build_completed_session(t_hdrs, math_id, [b"detect_alice", b"detect_alice"])

    resp = client.get(f"/attendance/session/{sess_id}", headers=t_hdrs)
    assert resp.status_code == 200, resp.text
    results = resp.json()

    # Should have 2 entries (one per student)
    assert len(results) == 2

    by_name = {r["student_name"]: r for r in results}
    assert "Alice" in by_name
    assert "Bob" in by_name

    alice_res = by_name["Alice"]
    bob_res = by_name["Bob"]

    assert alice_res["status"] == "present"
    assert bob_res["status"] == "absent"

    assert alice_res["total_scans"] == 2
    assert bob_res["total_scans"] == 2

    # Alice was detected in both scans
    assert alice_res["scans_detected"] == 2
    # Bob was not detected in any scan
    assert bob_res["scans_detected"] == 0


def test_session_attendance_404_if_not_found():
    _, t_token = register_and_login(
        {"name": "T404", "email": "t404@test.com", "password": "pass", "role": "teacher"}
    )
    resp = client.get("/attendance/session/99999", headers=auth_headers(t_token))
    assert resp.status_code == 404


def test_session_attendance_400_if_still_active():
    """Active session → 400 (attendance not computed yet)."""
    _, t_token = register_and_login(
        {"name": "TActive", "email": "tactive@test.com", "password": "pass", "role": "teacher"}
    )
    t_hdrs = auth_headers(t_token)
    subj_id = client.post("/subjects/", json={"name": "ActiveSub", "code": "ACT"}, headers=t_hdrs).json()["id"]
    sess_resp = client.post(
        "/sessions/start",
        json={"subject_id": subj_id, "scan_interval_seconds": 15},
        headers=t_hdrs,
    )
    sess_id = sess_resp.json()["session_id"]

    resp = client.get(f"/attendance/session/{sess_id}", headers=t_hdrs)
    assert resp.status_code == 400
    assert "still active" in resp.json()["detail"].lower()


def test_session_attendance_403_non_owner():
    """Non-owner teacher cannot view another teacher's session."""
    teacher1, t1_token = register_and_login(
        {"name": "T1", "email": "t1@test.com", "password": "pass", "role": "teacher"}
    )
    teacher2, t2_token = register_and_login(
        {"name": "T2", "email": "t2@test.com", "password": "pass", "role": "teacher"}
    )
    t1_hdrs = auth_headers(t1_token)
    t2_hdrs = auth_headers(t2_token)

    subj_id = client.post("/subjects/", json={"name": "Sub1", "code": "S1"}, headers=t1_hdrs).json()["id"]
    sess_id = build_completed_session(t1_hdrs, subj_id, [b"detect_none"])

    # Teacher 2 tries to access Teacher 1's session
    resp = client.get(f"/attendance/session/{sess_id}", headers=t2_hdrs)
    assert resp.status_code == 403


# ─────────────────────────────────────────────────────────────────────────────
# TEST 3 — Excel export: valid .xlsx, correct headers, correct row count
# ─────────────────────────────────────────────────────────────────────────────
def test_excel_export_valid_xlsx_headers_and_rows():
    """
    3 students in a completed session.
    Export should return a valid .xlsx with:
    - Sheet "Attendance Report"
    - Header: Student ID, Student Name, Subject, Date, Status, Scans Detected, Total Scans
    - 3 data rows (one per student)
    - Header row has blue fill
    - Present rows have green fill; Absent rows have red fill
    """
    _, t_token = register_and_login(
        {"name": "ExportTeacher", "email": "exportt@test.com", "password": "pass", "role": "teacher"}
    )
    t_hdrs = auth_headers(t_token)

    for i in range(1, 4):
        stu_data, _ = register_and_login(
            {"name": f"Student{i}", "email": f"stu{i}@test.com", "password": "pass",
             "role": "student", "student_id": f"STU_E{i}"}
        )
        student_mapping[stu_data["id"]] = f"Student{i}"

    subj_id = client.post(
        "/subjects/", json={"name": "Data Structures", "code": "CS301"}, headers=t_hdrs
    ).json()["id"]

    # detect_all → all 3 students present
    sess_id = build_completed_session(t_hdrs, subj_id, [b"detect_all", b"detect_all"])

    resp = client.get(f"/attendance/export/{sess_id}", headers=t_hdrs)
    assert resp.status_code == 200

    # Must be xlsx MIME type
    content_type = resp.headers.get("content-type", "")
    assert "spreadsheetml" in content_type, f"Unexpected Content-Type: {content_type}"

    # Must have Content-Disposition with filename
    disposition = resp.headers.get("content-disposition", "")
    assert f"attendance_{sess_id}.xlsx" in disposition

    # Load workbook from bytes and validate structure
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Attendance Report" in wb.sheetnames

    ws = wb["Attendance Report"]

    # ── Validate header row ────────────────────────────────────────────────
    expected_headers = [
        "Student ID", "Student Name", "Subject", "Date",
        "Status", "Scans Detected", "Total Scans"
    ]
    actual_headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
    assert actual_headers == expected_headers, f"Headers mismatch: {actual_headers}"

    # Header fill must be blue (3B5BDB).
    # openpyxl stores colors as 8-char ARGB (e.g. "FF3B5BDB"), so strip the
    # leading alpha channel before comparing.
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        raw_rgb = cell.fill.fgColor.rgb.upper()   # e.g. "FF3B5BDB"
        color_hex = raw_rgb[-6:]                  # last 6 chars = "3B5BDB"
        assert color_hex == "3B5BDB", \
            f"Expected blue header at col {col}, got fill={raw_rgb}"
        assert cell.font.bold is True
        # font color is also ARGB; white = FFFFFFFF or just FFFFFF
        font_rgb = cell.font.color.rgb.upper()
        assert font_rgb.endswith("FFFFFF"), \
            f"Expected white font at col {col}, got {font_rgb}"

    # ── Validate data rows ─────────────────────────────────────────────────
    data_rows = list(ws.iter_rows(min_row=2, values_only=False))
    assert len(data_rows) == 3, f"Expected 3 data rows (one per student), got {len(data_rows)}"

    for row_cells in data_rows:
        status_val = row_cells[4].value  # col 5 = Status
        assert status_val in ("Present", "Absent"), f"Unexpected status: {status_val}"

        # Green fill for present, red fill for absent.
        # Strip ARGB alpha prefix before comparing.
        raw_fill = row_cells[0].fill.fgColor.rgb.upper()
        fill_color = raw_fill[-6:]  # last 6 chars
        if status_val == "Present":
            assert fill_color == "C8F7C5", f"Expected green fill for Present, got {raw_fill}"
        else:
            assert fill_color == "FADADD", f"Expected red fill for Absent, got {raw_fill}"


# ─────────────────────────────────────────────────────────────────────────────
# TEST 4 — Export rejected if session still active
# ─────────────────────────────────────────────────────────────────────────────
def test_export_rejected_if_session_active():
    """Export should return 400 for an active (not yet stopped) session."""
    _, t_token = register_and_login(
        {"name": "TExp", "email": "texp@test.com", "password": "pass", "role": "teacher"}
    )
    t_hdrs = auth_headers(t_token)
    subj_id = client.post("/subjects/", json={"name": "TempSub", "code": "TMP"}, headers=t_hdrs).json()["id"]

    sess_resp = client.post(
        "/sessions/start",
        json={"subject_id": subj_id, "scan_interval_seconds": 15},
        headers=t_hdrs,
    )
    assert sess_resp.status_code == 200
    sess_id = sess_resp.json()["session_id"]

    resp = client.get(f"/attendance/export/{sess_id}", headers=t_hdrs)
    assert resp.status_code == 400
    assert "still active" in resp.json()["detail"].lower()


# ─────────────────────────────────────────────────────────────────────────────
# TEST 5 — Non-owner teacher cannot access another teacher's session export
# ─────────────────────────────────────────────────────────────────────────────
def test_export_403_for_non_owner_teacher():
    """Teacher 2 must get 403 when trying to export Teacher 1's session."""
    _, t1_token = register_and_login(
        {"name": "Owner", "email": "owner@test.com", "password": "pass", "role": "teacher"}
    )
    _, t2_token = register_and_login(
        {"name": "NonOwner", "email": "nonowner@test.com", "password": "pass", "role": "teacher"}
    )
    t1_hdrs = auth_headers(t1_token)
    t2_hdrs = auth_headers(t2_token)

    subj_id = client.post("/subjects/", json={"name": "OwnerSub", "code": "OWN"}, headers=t1_hdrs).json()["id"]
    sess_id = build_completed_session(t1_hdrs, subj_id, [b"detect_none"])

    # Teacher 2 tries to export Teacher 1's session → 403
    resp = client.get(f"/attendance/export/{sess_id}", headers=t2_hdrs)
    assert resp.status_code == 403
    assert "own" in resp.json()["detail"].lower()


if __name__ == "__main__":
    import sys
    sys.exit(pytest.main(["-v", __file__]))
